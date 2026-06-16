"""Parser de archivos de padrón (CSV / XLSX) — C-09."""

import csv
import io
from dataclasses import dataclass


class PadronParseError(ValueError):
    pass


@dataclass(frozen=True)
class FilaPadron:
    nombre: str
    apellidos: str
    email: str
    comision: str | None = None
    regional: str | None = None


_HEADER_ALIASES: dict[str, str] = {
    "nombre": "nombre",
    "name": "nombre",
    "first_name": "nombre",
    "apellido": "apellidos",
    "apellidos": "apellidos",
    "last_name": "apellidos",
    "email": "email",
    "correo": "email",
    "mail": "email",
    "comision": "comision",
    "grupo": "comision",
    "group": "comision",
    "regional": "regional",
    "sede": "regional",
}


def _normalize_header(value: str) -> str:
    return value.strip().lower().replace(" ", "_")


def _map_headers(raw_headers: list[str]) -> dict[int, str]:
    mapped: dict[int, str] = {}
    for idx, raw in enumerate(raw_headers):
        key = _normalize_header(raw)
        field = _HEADER_ALIASES.get(key)
        if field:
            mapped[idx] = field
    required = {"nombre", "apellidos", "email"}
    if not required.issubset(set(mapped.values())):
        raise PadronParseError(
            "El archivo debe incluir columnas nombre, apellidos y email"
        )
    return mapped


def parse_csv(content: bytes) -> list[FilaPadron]:
    text = content.decode("utf-8-sig")
    reader = csv.reader(io.StringIO(text))
    try:
        raw_headers = next(reader)
    except StopIteration as exc:
        raise PadronParseError("El archivo CSV está vacío") from exc

    header_map = _map_headers(raw_headers)
    filas: list[FilaPadron] = []
    for row in reader:
        if not any(cell.strip() for cell in row):
            continue
        data: dict[str, str] = {}
        for idx, field in header_map.items():
            if idx < len(row):
                data[field] = row[idx].strip()
        if not data.get("email"):
            continue
        filas.append(
            FilaPadron(
                nombre=data.get("nombre", ""),
                apellidos=data.get("apellidos", ""),
                email=data["email"],
                comision=data.get("comision") or None,
                regional=data.get("regional") or None,
            )
        )
    if not filas:
        raise PadronParseError("No se encontraron filas válidas en el archivo")
    return filas


def parse_xlsx(content: bytes) -> list[FilaPadron]:
    try:
        from openpyxl import load_workbook
    except ImportError as exc:
        raise PadronParseError(
            "Soporte XLSX no disponible (falta openpyxl)"
        ) from exc

    wb = load_workbook(io.BytesIO(content), read_only=True, data_only=True)
    sheet = wb.active
    rows = sheet.iter_rows(values_only=True)
    try:
        raw_headers = [str(c or "") for c in next(rows)]
    except StopIteration as exc:
        raise PadronParseError("El archivo XLSX está vacío") from exc

    header_map = _map_headers(raw_headers)
    filas: list[FilaPadron] = []
    for row in rows:
        if row is None or not any(row):
            continue
        cells = [str(c or "").strip() for c in row]
        data: dict[str, str] = {}
        for idx, field in header_map.items():
            if idx < len(cells):
                data[field] = cells[idx]
        if not data.get("email"):
            continue
        filas.append(
            FilaPadron(
                nombre=data.get("nombre", ""),
                apellidos=data.get("apellidos", ""),
                email=data["email"],
                comision=data.get("comision") or None,
                regional=data.get("regional") or None,
            )
        )
    if not filas:
        raise PadronParseError("No se encontraron filas válidas en el archivo")
    return filas


def parse_padron_file(content: bytes, filename: str) -> list[FilaPadron]:
    lower = filename.lower()
    if lower.endswith(".csv"):
        return parse_csv(content)
    if lower.endswith(".xlsx"):
        return parse_xlsx(content)
    raise PadronParseError("Formato no soportado. Use .csv o .xlsx")
